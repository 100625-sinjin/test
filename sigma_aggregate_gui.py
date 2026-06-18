#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
シグマ集計Excel CSV転記ツール

添付形式の集計Excelファイルを読み込み、品種ごとに指定したCSVフォルダから
日付別の実績を集計Excelへ転記するGUIアプリケーションです。

必要ライブラリ:
  - Python 3.10 以降推奨
  - openpyxl
  - tkinter（通常のPythonには同梱）

起動:
  python sigma_aggregate_gui.py
"""
from __future__ import annotations

import csv
import datetime as dt
import os
import re
import sys
import traceback
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple, Any

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    from tkinter import ttk
except Exception as exc:  # pragma: no cover
    raise RuntimeError("tkinter を読み込めません。Python の tkinter 付き環境で実行してください。") from exc

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except Exception as exc:  # pragma: no cover
    raise RuntimeError("openpyxl が必要です。`pip install openpyxl` を実行してください。") from exc

APP_TITLE = "シグマ集計 Excel CSV転記ツール"
SUPPORTED_EXT = ".csv"
ITEM_BOARD_COUNT = "生産基板枚数"
ITEM_OPERATION_TIME = "生産運転時間"
ITEM_MOUNT_TIME = "装着時間"
ITEM_BOARD_WAIT_TIME = "基板待ち時間"
ITEMS_TO_WRITE = [ITEM_BOARD_COUNT, ITEM_OPERATION_TIME, ITEM_MOUNT_TIME, ITEM_BOARD_WAIT_TIME]


def normalize_text(value: Any) -> str:
    """検索・照合用に文字列を正規化する。"""
    if value is None:
        return ""
    s = str(value)
    s = s.replace("　", " ")
    s = re.sub(r"\s+", "", s)
    s = s.replace("\\_", "_")
    return s.strip()


def parse_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def parse_duration_to_seconds(value: Any) -> Optional[int]:
    """H:MM:SS / HH:MM:SS / 0:00'00 形式を秒へ変換する。"""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s == "-":
        return None
    s = s.replace("'", ":")
    m = re.fullmatch(r"(\d+):(\d{1,2}):(\d{1,2})", s)
    if not m:
        return None
    hours, minutes, seconds = map(int, m.groups())
    if minutes >= 60 or seconds >= 60:
        return None
    return hours * 3600 + minutes * 60 + seconds


def seconds_to_excel_days(seconds: Optional[int]) -> Optional[float]:
    if seconds is None:
        return None
    return seconds / 86400.0


def seconds_to_hms(seconds: Optional[int]) -> str:
    if seconds is None:
        return ""
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    return f"{h}:{m:02d}:{s:02d}"


def parse_japanese_datetime(value: str) -> Optional[dt.datetime]:
    """例: 2026年 6月1日(月) 16:53"""
    if not value:
        return None
    s = str(value).strip()
    m = re.search(r"(20\d{2})年\s*(\d{1,2})月\s*(\d{1,2})日(?:\([^)]*\))?\s*(\d{1,2}):(\d{2})", s)
    if not m:
        return None
    y, mo, d, hh, mm = map(int, m.groups())
    try:
        return dt.datetime(y, mo, d, hh, mm)
    except ValueError:
        return None


def parse_date_from_filename(path: Path) -> Optional[dt.date]:
    """ファイル名から日付を抽出する。YYYY-MM-DD / YYYY_MM_DD / YYYYMMDD に対応。"""
    name = path.name
    patterns = [
        r"(?P<y>20\d{2})[-_\. ](?P<m>\d{1,2})[-_\. ](?P<d>\d{1,2})",
        r"(?P<y>20\d{2})(?P<m>\d{2})(?P<d>\d{2})",
    ]
    for pat in patterns:
        m = re.search(pat, name)
        if m:
            try:
                return dt.date(int(m.group("y")), int(m.group("m")), int(m.group("d")))
            except ValueError:
                return None
    return None


def read_csv_rows(path: Path) -> List[List[str]]:
    """CSVを複数の日本語エンコーディング候補で読み込む。"""
    encodings = ["utf-8-sig", "cp932", "shift_jis", "utf-8"]
    last_error: Optional[Exception] = None
    for enc in encodings:
        try:
            with path.open("r", encoding=enc, newline="") as f:
                return [row for row in csv.reader(f)]
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    raise UnicodeDecodeError("csv", b"", 0, 1, f"CSVの文字コードを判別できません: {path} / {last_error}")


@dataclass
class CsvMetrics:
    path: Path
    date: Optional[dt.date]
    internal_date: Optional[dt.date]
    board_count: Optional[int]
    operation_seconds: Optional[int]
    mount_seconds: Optional[int]
    wait_before_seconds: Optional[int]
    wait_after_seconds: Optional[int]
    product_comment: str = ""

    @property
    def wait_total_seconds(self) -> Optional[int]:
        if self.wait_before_seconds is None and self.wait_after_seconds is None:
            return None
        return (self.wait_before_seconds or 0) + (self.wait_after_seconds or 0)


def parse_csv_metrics(path: Path, prefer_filename_date: bool = True) -> CsvMetrics:
    rows = read_csv_rows(path)

    board_count: Optional[int] = None
    operation_seconds: Optional[int] = None
    mount_seconds: Optional[int] = None
    wait_before_seconds: Optional[int] = None
    wait_after_seconds: Optional[int] = None
    internal_date: Optional[dt.date] = None
    product_comment = ""

    in_machine_info = False
    machine_header_seen = False

    for row in rows:
        cells = [c.strip() for c in row]
        first = cells[0] if cells else ""
        key = normalize_text(first)

        if internal_date is None:
            for c in cells[:2]:
                parsed_dt = parse_japanese_datetime(c)
                if parsed_dt:
                    internal_date = parsed_dt.date()
                    break

        if key == normalize_text("[機種別稼働情報]"):
            in_machine_info = True
            machine_header_seen = False
            continue
        if in_machine_info and not machine_header_seen:
            machine_header_seen = True
            continue
        if in_machine_info and machine_header_seen and cells and cells[0]:
            # パターンプログラム名,コメント,更新日時,装置名 の2列目が品種コメント
            if len(cells) >= 2:
                product_comment = cells[1].strip()
            in_machine_info = False

        if key == normalize_text(ITEM_BOARD_COUNT) and board_count is None:
            board_count = parse_int(cells[1] if len(cells) > 1 else None)
            continue
        if key == normalize_text(ITEM_OPERATION_TIME) and operation_seconds is None:
            operation_seconds = parse_duration_to_seconds(cells[1] if len(cells) > 1 else None)
            continue
        if key == normalize_text(ITEM_MOUNT_TIME) and mount_seconds is None:
            mount_seconds = parse_duration_to_seconds(cells[1] if len(cells) > 1 else None)
            continue
        if key == normalize_text("基板待ち時間前工程待ち") and wait_before_seconds is None:
            wait_before_seconds = parse_duration_to_seconds(cells[1] if len(cells) > 1 else None)
            continue
        if key == normalize_text("後工程待ち") and wait_after_seconds is None:
            wait_after_seconds = parse_duration_to_seconds(cells[1] if len(cells) > 1 else None)
            continue

    filename_date = parse_date_from_filename(path)
    if prefer_filename_date:
        use_date = filename_date or internal_date
    else:
        use_date = internal_date or filename_date

    return CsvMetrics(
        path=path,
        date=use_date,
        internal_date=internal_date,
        board_count=board_count,
        operation_seconds=operation_seconds,
        mount_seconds=mount_seconds,
        wait_before_seconds=wait_before_seconds,
        wait_after_seconds=wait_after_seconds,
        product_comment=product_comment,
    )


@dataclass
class ProductBlock:
    name: str
    start_row: int
    item_rows: Dict[str, int]


def cell_date_value(value: Any) -> Optional[dt.date]:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return None


def build_date_columns(ws) -> Dict[dt.date, int]:
    """3行目のD列以降を日付列として認識する。数式列は左列+1日として推定する。"""
    date_cols: Dict[dt.date, int] = {}
    current_date: Optional[dt.date] = None
    for col in range(4, ws.max_column + 1):
        value = ws.cell(3, col).value
        parsed = cell_date_value(value)
        if parsed is not None:
            current_date = parsed
        elif isinstance(value, str) and value.startswith("=") and current_date is not None:
            current_date = current_date + dt.timedelta(days=1)
        elif current_date is not None:
            # 空欄があっても、帳票が連続日付形式なら日付を進める
            current_date = current_date + dt.timedelta(days=1)
        else:
            continue
        date_cols[current_date] = col
    return date_cols


def find_product_blocks(ws) -> List[ProductBlock]:
    blocks: List[ProductBlock] = []
    for row in range(1, ws.max_row + 1):
        product = ws.cell(row, 2).value
        first_item = ws.cell(row, 3).value
        if not product or normalize_text(first_item) != normalize_text(ITEM_BOARD_COUNT):
            continue
        item_rows: Dict[str, int] = {}
        r = row
        while r <= ws.max_row:
            if r > row and ws.cell(r, 2).value:
                break
            item = ws.cell(r, 3).value
            if item:
                item_rows[normalize_text(item)] = r
            r += 1
        blocks.append(ProductBlock(str(product).strip(), row, item_rows))
    return blocks


def flexible_product_match(product_name: str, comment: str) -> bool:
    """CSVコメントがExcel品種と明らかに矛盾しないかを緩く確認する。"""
    if not comment:
        return True
    p = normalize_text(product_name).upper()
    c = normalize_text(comment).upper()
    if p in c:
        return True
    # 例: ZG-18(A)_MP と ZG-18(A)_FPC387_MP_Winbond を同一系統として扱う
    tokens = [t for t in re.split(r"[_\-]", p) if t]
    return bool(tokens and tokens[0] in c)


def write_metrics_to_sheet(ws, block: ProductBlock, date_col: int, metrics: CsvMetrics, overwrite: bool = True) -> List[str]:
    messages: List[str] = []
    write_map: Dict[str, Any] = {
        ITEM_BOARD_COUNT: metrics.board_count,
        ITEM_OPERATION_TIME: seconds_to_excel_days(metrics.operation_seconds),
        ITEM_MOUNT_TIME: seconds_to_excel_days(metrics.mount_seconds),
        ITEM_BOARD_WAIT_TIME: seconds_to_excel_days(metrics.wait_total_seconds),
    }
    for item_name, value in write_map.items():
        row = block.item_rows.get(normalize_text(item_name))
        if row is None:
            messages.append(f"項目なし: {block.name} / {item_name}")
            continue
        if value is None:
            messages.append(f"CSV値なし: {metrics.path.name} / {item_name}")
            continue
        cell = ws.cell(row, date_col)
        if not overwrite and cell.value not in (None, ""):
            messages.append(f"既存値のためスキップ: {block.name} / {ws.cell(3, date_col).coordinate} / {item_name}")
            continue
        cell.value = value
        if item_name == ITEM_BOARD_COUNT:
            cell.number_format = "0"
        else:
            # 24時間超にも対応するため [h]:mm:ss を使用
            cell.number_format = "[h]:mm:ss"
    return messages


@dataclass
class PreviewRecord:
    product: str
    csv_path: Path
    date: Optional[dt.date]
    metrics: CsvMetrics
    status: str


class SigmaAggregateApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1180x760")
        self.minsize(1060, 650)

        self.excel_path_var = tk.StringVar()
        self.output_path_var = tk.StringVar()
        self.prefer_filename_date_var = tk.BooleanVar(value=True)
        self.overwrite_var = tk.BooleanVar(value=True)

        self.product_blocks: List[ProductBlock] = []
        self.date_cols: Dict[dt.date, int] = {}
        self.product_folders: Dict[str, str] = {}
        self.preview_records: List[PreviewRecord] = []

        self._setup_style()
        self._build_ui()

    def _setup_style(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        base_font = ("Yu Gothic UI", 10)
        self.option_add("*Font", base_font)
        style.configure("Title.TLabel", font=("Yu Gothic UI", 16, "bold"), padding=(0, 4))
        style.configure("Sub.TLabel", foreground="#555555")
        style.configure("Accent.TButton", font=("Yu Gothic UI", 10, "bold"))
        style.configure("Treeview", rowheight=26)
        style.configure("Treeview.Heading", font=("Yu Gothic UI", 10, "bold"))

    def _build_ui(self) -> None:
        root = ttk.Frame(self, padding=14)
        root.pack(fill=tk.BOTH, expand=True)

        header = ttk.Frame(root)
        header.pack(fill=tk.X)
        ttk.Label(header, text=APP_TITLE, style="Title.TLabel").pack(anchor=tk.W)
        ttk.Label(
            header,
            text="集計Excelの品種別ブロックへ、CSVフォルダ内の日付別データを安全に転記します。",
            style="Sub.TLabel",
        ).pack(anchor=tk.W)

        file_frame = ttk.LabelFrame(root, text="1. 集計Excelと出力先", padding=10)
        file_frame.pack(fill=tk.X, pady=(12, 8))
        file_frame.columnconfigure(1, weight=1)
        ttk.Label(file_frame, text="集計Excel").grid(row=0, column=0, sticky=tk.W, padx=(0, 8), pady=3)
        ttk.Entry(file_frame, textvariable=self.excel_path_var).grid(row=0, column=1, sticky=tk.EW, pady=3)
        ttk.Button(file_frame, text="参照", command=self.browse_excel).grid(row=0, column=2, padx=(8, 0), pady=3)
        ttk.Button(file_frame, text="読み込み", style="Accent.TButton", command=self.load_excel).grid(row=0, column=3, padx=(8, 0), pady=3)

        ttk.Label(file_frame, text="出力Excel").grid(row=1, column=0, sticky=tk.W, padx=(0, 8), pady=3)
        ttk.Entry(file_frame, textvariable=self.output_path_var).grid(row=1, column=1, sticky=tk.EW, pady=3)
        ttk.Button(file_frame, text="保存先", command=self.browse_output).grid(row=1, column=2, padx=(8, 0), pady=3)

        options = ttk.Frame(file_frame)
        options.grid(row=2, column=1, columnspan=3, sticky=tk.W, pady=(6, 0))
        ttk.Checkbutton(options, text="CSV日付はファイル名を優先する", variable=self.prefer_filename_date_var).pack(side=tk.LEFT, padx=(0, 18))
        ttk.Checkbutton(options, text="既存値を上書きする", variable=self.overwrite_var).pack(side=tk.LEFT)

        middle = ttk.Frame(root)
        middle.pack(fill=tk.BOTH, expand=True)
        middle.columnconfigure(0, weight=4)
        middle.columnconfigure(1, weight=6)
        middle.rowconfigure(0, weight=1)

        left = ttk.LabelFrame(middle, text="2. 品種ごとのCSVフォルダ", padding=10)
        left.grid(row=0, column=0, sticky=tk.NSEW, padx=(0, 8))
        left.rowconfigure(0, weight=1)
        left.columnconfigure(0, weight=1)
        self.product_tree = ttk.Treeview(left, columns=("product", "folder"), show="headings", selectmode="browse")
        self.product_tree.heading("product", text="品種")
        self.product_tree.heading("folder", text="CSVフォルダ")
        self.product_tree.column("product", width=150, anchor=tk.W)
        self.product_tree.column("folder", width=360, anchor=tk.W)
        self.product_tree.grid(row=0, column=0, sticky=tk.NSEW)
        product_scroll = ttk.Scrollbar(left, orient=tk.VERTICAL, command=self.product_tree.yview)
        product_scroll.grid(row=0, column=1, sticky=tk.NS)
        self.product_tree.configure(yscrollcommand=product_scroll.set)

        left_buttons = ttk.Frame(left)
        left_buttons.grid(row=1, column=0, columnspan=2, sticky=tk.EW, pady=(8, 0))
        ttk.Button(left_buttons, text="選択品種のフォルダ指定", command=self.set_folder_for_selected).pack(side=tk.LEFT)
        ttk.Button(left_buttons, text="全品種に同じフォルダ", command=self.set_folder_for_all).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(left_buttons, text="フォルダ解除", command=self.clear_selected_folder).pack(side=tk.LEFT, padx=(8, 0))

        right = ttk.LabelFrame(middle, text="3. 読み取りプレビュー", padding=10)
        right.grid(row=0, column=1, sticky=tk.NSEW)
        right.rowconfigure(0, weight=1)
        right.columnconfigure(0, weight=1)
        columns = ("product", "date", "csv", "boards", "operation", "mount", "wait", "status")
        self.preview_tree = ttk.Treeview(right, columns=columns, show="headings")
        headings = {
            "product": "品種",
            "date": "日付",
            "csv": "CSV",
            "boards": "生産基板枚数",
            "operation": "生産運転時間",
            "mount": "装着時間",
            "wait": "基板待ち時間",
            "status": "状態",
        }
        widths = {"product": 120, "date": 92, "csv": 230, "boards": 90, "operation": 92, "mount": 82, "wait": 92, "status": 160}
        for col in columns:
            self.preview_tree.heading(col, text=headings[col])
            self.preview_tree.column(col, width=widths[col], anchor=tk.W)
        self.preview_tree.grid(row=0, column=0, sticky=tk.NSEW)
        preview_scroll = ttk.Scrollbar(right, orient=tk.VERTICAL, command=self.preview_tree.yview)
        preview_scroll.grid(row=0, column=1, sticky=tk.NS)
        self.preview_tree.configure(yscrollcommand=preview_scroll.set)

        action_frame = ttk.Frame(root)
        action_frame.pack(fill=tk.X, pady=(10, 8))
        ttk.Button(action_frame, text="CSVを読み取りプレビュー", style="Accent.TButton", command=self.scan_csv).pack(side=tk.LEFT)
        ttk.Button(action_frame, text="Excelへ転記して保存", style="Accent.TButton", command=self.export_excel).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(action_frame, text="終了", command=self.destroy).pack(side=tk.RIGHT)

        log_frame = ttk.LabelFrame(root, text="ログ", padding=8)
        log_frame.pack(fill=tk.BOTH, expand=False)
        self.log_text = tk.Text(log_frame, height=7, wrap=tk.WORD)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log("準備完了。まず集計Excelを選択して読み込んでください。")

    def log(self, message: str) -> None:
        timestamp = dt.datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.update_idletasks()

    def browse_excel(self) -> None:
        path = filedialog.askopenfilename(
            title="集計Excelを選択",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
        )
        if path:
            self.excel_path_var.set(path)
            p = Path(path)
            default_output = p.with_name(f"{p.stem}_集計済_{dt.datetime.now():%Y%m%d_%H%M%S}{p.suffix}")
            self.output_path_var.set(str(default_output))

    def browse_output(self) -> None:
        initial = self.output_path_var.get() or "集計済.xlsx"
        path = filedialog.asksaveasfilename(
            title="出力Excelの保存先",
            defaultextension=".xlsx",
            initialfile=Path(initial).name,
            initialdir=str(Path(initial).parent) if Path(initial).parent.exists() else os.getcwd(),
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.output_path_var.set(path)

    def load_excel(self) -> None:
        try:
            path = Path(self.excel_path_var.get())
            if not path.exists():
                messagebox.showerror("エラー", "集計Excelが見つかりません。")
                return
            wb = load_workbook(path, data_only=False)
            ws = wb.active
            self.product_blocks = find_product_blocks(ws)
            self.date_cols = build_date_columns(ws)
            if not self.product_blocks:
                messagebox.showerror("エラー", "品種ブロックを検出できません。B列=品種、C列=生産基板枚数 の形式を確認してください。")
                return
            if not self.date_cols:
                messagebox.showerror("エラー", "日付列を検出できません。3行目D列以降の日付を確認してください。")
                return
            self.product_folders = {block.name: self.product_folders.get(block.name, "") for block in self.product_blocks}
            self.refresh_product_tree()
            dates = sorted(self.date_cols)
            self.log(f"Excel読込完了: 品種 {len(self.product_blocks)}件 / 日付 {dates[0]} ～ {dates[-1]} ({len(dates)}列)")
        except Exception as exc:
            self.log(traceback.format_exc())
            messagebox.showerror("エラー", f"Excel読込に失敗しました。\n{exc}")

    def refresh_product_tree(self) -> None:
        for item in self.product_tree.get_children():
            self.product_tree.delete(item)
        for block in self.product_blocks:
            self.product_tree.insert("", tk.END, values=(block.name, self.product_folders.get(block.name, "")))

    def selected_product_name(self) -> Optional[str]:
        sel = self.product_tree.selection()
        if not sel:
            return None
        values = self.product_tree.item(sel[0], "values")
        return values[0] if values else None

    def set_folder_for_selected(self) -> None:
        product = self.selected_product_name()
        if not product:
            messagebox.showinfo("情報", "品種を選択してください。")
            return
        folder = filedialog.askdirectory(title=f"{product} のCSVフォルダを選択")
        if folder:
            self.product_folders[product] = folder
            self.refresh_product_tree()
            self.log(f"フォルダ指定: {product} -> {folder}")

    def set_folder_for_all(self) -> None:
        if not self.product_blocks:
            messagebox.showinfo("情報", "先に集計Excelを読み込んでください。")
            return
        folder = filedialog.askdirectory(title="全品種に設定するCSVフォルダを選択")
        if folder:
            for block in self.product_blocks:
                self.product_folders[block.name] = folder
            self.refresh_product_tree()
            self.log(f"全品種に同じフォルダを指定: {folder}")

    def clear_selected_folder(self) -> None:
        product = self.selected_product_name()
        if not product:
            messagebox.showinfo("情報", "品種を選択してください。")
            return
        self.product_folders[product] = ""
        self.refresh_product_tree()
        self.log(f"フォルダ解除: {product}")

    def scan_csv(self) -> None:
        if not self.product_blocks or not self.date_cols:
            messagebox.showinfo("情報", "先に集計Excelを読み込んでください。")
            return
        self.preview_records.clear()
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)

        prefer_filename = self.prefer_filename_date_var.get()
        total_files = 0
        for block in self.product_blocks:
            folder = self.product_folders.get(block.name, "")
            if not folder:
                continue
            folder_path = Path(folder)
            if not folder_path.exists():
                self.log(f"フォルダなし: {block.name} / {folder}")
                continue
            csv_files = sorted(folder_path.glob(f"*{SUPPORTED_EXT}"))
            for csv_path in csv_files:
                total_files += 1
                try:
                    metrics = parse_csv_metrics(csv_path, prefer_filename_date=prefer_filename)
                    status_parts: List[str] = []
                    if metrics.date is None:
                        status_parts.append("日付なし")
                    elif metrics.date not in self.date_cols:
                        status_parts.append("Excel日付範囲外")
                    if not flexible_product_match(block.name, metrics.product_comment):
                        status_parts.append("品種コメント要確認")
                    missing = []
                    if metrics.board_count is None:
                        missing.append(ITEM_BOARD_COUNT)
                    if metrics.operation_seconds is None:
                        missing.append(ITEM_OPERATION_TIME)
                    if metrics.mount_seconds is None:
                        missing.append(ITEM_MOUNT_TIME)
                    if metrics.wait_total_seconds is None:
                        missing.append(ITEM_BOARD_WAIT_TIME)
                    if missing:
                        status_parts.append("値なし: " + ",".join(missing))
                    status = "OK" if not status_parts else " / ".join(status_parts)
                    rec = PreviewRecord(block.name, csv_path, metrics.date, metrics, status)
                    self.preview_records.append(rec)
                    self.preview_tree.insert("", tk.END, values=(
                        block.name,
                        metrics.date.isoformat() if metrics.date else "",
                        csv_path.name,
                        metrics.board_count if metrics.board_count is not None else "",
                        seconds_to_hms(metrics.operation_seconds),
                        seconds_to_hms(metrics.mount_seconds),
                        seconds_to_hms(metrics.wait_total_seconds),
                        status,
                    ))
                except Exception as exc:
                    self.log(f"CSV読込エラー: {csv_path.name} / {exc}")
        self.log(f"CSVプレビュー完了: {len(self.preview_records)}件（検出CSV {total_files}件）")
        if not self.preview_records:
            messagebox.showinfo("情報", "CSVが見つかりません。品種ごとのフォルダ指定を確認してください。")

    def export_excel(self) -> None:
        if not self.preview_records:
            self.scan_csv()
            if not self.preview_records:
                return
        excel_path = Path(self.excel_path_var.get())
        output_path = Path(self.output_path_var.get()) if self.output_path_var.get() else None
        if not excel_path.exists():
            messagebox.showerror("エラー", "集計Excelが見つかりません。")
            return
        if output_path is None:
            messagebox.showerror("エラー", "出力Excelを指定してください。")
            return
        try:
            wb = load_workbook(excel_path, data_only=False)
            ws = wb.active
            blocks = {b.name: b for b in find_product_blocks(ws)}
            date_cols = build_date_columns(ws)
            overwrite = self.overwrite_var.get()

            written = 0
            skipped = 0
            warnings: List[str] = []
            seen_targets: set[Tuple[str, dt.date]] = set()

            for rec in self.preview_records:
                if rec.date is None or rec.date not in date_cols:
                    skipped += 1
                    continue
                block = blocks.get(rec.product)
                if block is None:
                    warnings.append(f"品種ブロックなし: {rec.product}")
                    skipped += 1
                    continue
                target_key = (rec.product, rec.date)
                if target_key in seen_targets:
                    warnings.append(f"同一品種・同一日付のCSVが複数あります。後続CSVで上書き: {rec.product} / {rec.date}")
                seen_targets.add(target_key)
                msg = write_metrics_to_sheet(ws, block, date_cols[rec.date], rec.metrics, overwrite=overwrite)
                warnings.extend(msg)
                if "OK" in rec.status or not any(s in rec.status for s in ["日付なし", "Excel日付範囲外", "値なし"]):
                    written += 1
                else:
                    # 書ける値だけは書き込まれている可能性がある
                    written += 1

            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            self.log(f"保存完了: {output_path}")
            self.log(f"転記対象CSV: {written}件 / スキップ: {skipped}件 / 警告: {len(warnings)}件")
            for w in warnings[:50]:
                self.log("警告: " + w)
            if len(warnings) > 50:
                self.log(f"警告が多いため先頭50件のみ表示しました。残り {len(warnings)-50}件")
            messagebox.showinfo("完了", f"Excelへ転記して保存しました。\n{output_path}")
        except Exception as exc:
            self.log(traceback.format_exc())
            messagebox.showerror("エラー", f"Excel保存に失敗しました。\n{exc}")


def main() -> None:
    app = SigmaAggregateApp()
    # コマンドライン引数にExcelパスが指定された場合は初期入力する
    if len(sys.argv) >= 2:
        p = Path(sys.argv[1])
        if p.exists():
            app.excel_path_var.set(str(p))
            app.output_path_var.set(str(p.with_name(f"{p.stem}_集計済_{dt.datetime.now():%Y%m%d_%H%M%S}{p.suffix}")))
    app.mainloop()


if __name__ == "__main__":
    main()
